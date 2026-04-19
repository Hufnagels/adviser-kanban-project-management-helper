import io
import re
from typing import Optional

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

try:
    from fpdf import FPDF
    _FPDF_AVAILABLE = True
except ImportError:
    _FPDF_AVAILABLE = False

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.task import Priority, Task, TaskStatus
from app.models.user import User

router = APIRouter(prefix="/import-export", tags=["import-export"])

TASK_COLUMNS = ["title", "description", "status", "priority", "listing_date", "due_date", "notes"]

DATE_COLUMNS = {"Listing Date", "Due Date", "Finishing Date"}
STATUS_FILLS = {
    "todo":        "D9D9D9",
    "in_progress": "BDD7EE",
    "in_review":   "FFEB9C",
    "done":        "C6EFCE",
}
PRIORITY_FILLS = {
    "low":      "D9D9D9",
    "medium":   "BDD7EE",
    "high":     "FCEBC4",
    "critical": "FFC7CE",
}

_thin = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
_DATA_ALIGN = Alignment(vertical="center", wrap_text=False)
_DATA_FONT = Font(name="Calibri", size=10)


def _style_worksheet(ws) -> None:
    """Apply professional formatting to the Tasks worksheet."""
    headers = [cell.value for cell in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers)}

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER

    # ── Freeze pane + auto-filter ─────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Column widths (auto-fit, capped at 42) ───────────────────────────────
    for col_cells in ws.iter_cols():
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 42)

    # ── Data rows ────────────────────────────────────────────────────────────
    status_col = col_index.get("Status")
    priority_col = col_index.get("Priority")
    notes_col = col_index.get("Notes")
    date_cols = {col_index[h] for h in DATE_COLUMNS if h in col_index}

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            is_notes = (cell.column == notes_col)
            cell.font = _DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=is_notes)
            cell.border = _BORDER

            if cell.column in date_cols:
                cell.number_format = "YYYY-MM-DD"

            if cell.column == status_col and cell.value:
                fill_hex = STATUS_FILLS.get(str(cell.value).lower())
                if fill_hex:
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)

            if cell.column == priority_col and cell.value:
                fill_hex = PRIORITY_FILLS.get(str(cell.value).lower())
                if fill_hex:
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)

        # Auto-height: estimate lines needed in Notes cell
        if notes_col:
            notes_val = str(row[notes_col - 1].value or "")
            notes_w = ws.column_dimensions[get_column_letter(notes_col)].width or 42
            chars_per_line = max(1, int(notes_w * 1.1))
            n_lines = max(1, -(-len(notes_val) // chars_per_line))
            ws.row_dimensions[row[0].row].height = max(15, n_lines * 14)
        else:
            ws.row_dimensions[row[0].row].height = 15


@router.post("/upload")
async def import_tasks(
    file: UploadFile = File(...),
    project_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")

    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")

    # Normalize column names but keep original mapping for "Task ID"
    col_map = {c: c.strip().lower().replace(" ", "_") for c in df.columns}
    df.rename(columns=col_map, inplace=True)

    if "title" not in df.columns:
        raise HTTPException(status_code=400, detail="Excel must have a 'title' column")

    # Drop rows where ALL cells are empty/NaN
    df.dropna(how="all", inplace=True)

    def safe_str(val) -> str | None:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        s = str(val).strip()
        return s if s and s.lower() not in ("nan", "none") else None

    created = 0
    skipped = 0
    for _, row in df.iterrows():
        title = safe_str(row.get("title"))
        if not title:
            continue

        # Preserve original Task ID from Excel (column "task_id")
        external_id = safe_str(row.get("task_id"))

        # ── Deduplication ─────────────────────────────────────────────────────
        # If external_id present: skip if same external_id + project already exists
        # Otherwise: skip if same title + project already exists
        if external_id:
            dup_q = select(Task).where(Task.external_id == external_id)
            if project_id:
                dup_q = dup_q.where(Task.project_id == project_id)
        else:
            dup_q = select(Task).where(Task.title == title)
            if project_id:
                dup_q = dup_q.where(Task.project_id == project_id)

        existing = await db.execute(dup_q.limit(1))
        if existing.scalar_one_or_none():
            skipped += 1
            continue
        # ─────────────────────────────────────────────────────────────────────

        status_raw = (safe_str(row.get("status")) or "todo").lower().replace(" ", "_")
        priority_raw = (safe_str(row.get("priority")) or "medium").lower()

        task = Task(
            external_id=external_id,
            title=title,
            task_type=safe_str(row.get("type")),
            description=safe_str(row.get("description")),
            status=TaskStatus(status_raw) if status_raw in TaskStatus._value2member_map_ else TaskStatus.todo,
            priority=Priority(priority_raw) if priority_raw in Priority._value2member_map_ else Priority.medium,
            notes=safe_str(row.get("notes")),
            project_id=project_id,
            created_by=current_user.id,
        )

        for date_field in ("listing_date", "due_date", "finishing_date"):
            val = row.get(date_field)
            if val is not None and pd.notna(val):
                try:
                    setattr(task, date_field, pd.to_datetime(val).date())
                except Exception:
                    pass

        db.add(task)
        created += 1

    await db.commit()
    return {"imported": created, "skipped": skipped}


_MULTI_HEADERS = ["Task ID", "Title", "Type", "Status", "Priority",
                   "Listing Date", "Due Date", "Finishing Date", "Notes", "Approval"]
_MULTI_COL_WIDTHS = [14, 42, 16, 16, 14, 14, 14, 16, 48, 10]
_MULTI_DATE_COLS  = {"Listing Date", "Due Date", "Finishing Date"}


def _task_to_row(t: Task) -> list:
    return [
        t.external_id or "",
        t.title,
        t.task_type or "",
        t.status.value,
        t.priority.value,
        t.listing_date,
        t.due_date,
        t.finishing_date,
        _html_to_text(t.notes),
        "Yes" if t.approval else "No",
    ]


def _write_multi_sheet(ws, tasks: list[Task]) -> None:
    """Write tasks to a worksheet with header row + data rows, styled."""
    # Column widths
    for ci, w in enumerate(_MULTI_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Header
    for ci, hdr in enumerate(_MULTI_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    col_idx = {h: i + 1 for i, h in enumerate(_MULTI_HEADERS)}
    status_col   = col_idx["Status"]
    priority_col = col_idx["Priority"]
    date_cols    = {col_idx[h] for h in _MULTI_DATE_COLS}

    notes_col_idx = col_idx["Notes"]
    notes_w = _MULTI_COL_WIDTHS[notes_col_idx - 1]
    chars_per_line = max(1, int(notes_w * 1.1))

    for ri, t in enumerate(tasks, start=2):
        vals = _task_to_row(t)
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = _DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == notes_col_idx))
            cell.border = _BORDER
            if ci in date_cols and val:
                cell.number_format = "YYYY-MM-DD"
            if ci == status_col and val:
                fhex = STATUS_FILLS.get(str(val).lower())
                if fhex:
                    cell.fill = PatternFill(fill_type="solid", fgColor=fhex)
            if ci == priority_col and val:
                fhex = PRIORITY_FILLS.get(str(val).lower())
                if fhex:
                    cell.fill = PatternFill(fill_type="solid", fgColor=fhex)
        notes_val = str(vals[notes_col_idx - 1] or "")
        n_lines = max(1, -(-len(notes_val) // chars_per_line))
        ws.row_dimensions[ri].height = max(15, n_lines * 14)


@router.get("/export-multi")
async def export_tasks_multi(
    project_ids: list[str] = Query(...),
    project_names: list[str] = Query(default=[]),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export tasks for multiple projects — one sheet per project + combined 'All Tasks' sheet."""
    from datetime import datetime as dt

    # Collect tasks per project
    project_tasks: list[tuple[str, list[Task]]] = []
    for i, pid in enumerate(project_ids):
        name = (project_names[i] if i < len(project_names) else f"Project {i + 1}")[:31]
        for ch in r'\/*?[]':
            name = name.replace(ch, "_")
        result = await db.execute(
            select(Task).where(Task.project_id == pid).order_by(Task.created_at.desc())
        )
        project_tasks.append((name, list(result.scalars().all())))

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── "All Tasks" combined sheet (first tab) ────────────────────────────────
    ws_all = wb.create_sheet("All Tasks")
    ws_all.sheet_properties.tabColor = "1F4E79"
    for ci, w in enumerate(_MULTI_COL_WIDTHS, start=1):
        ws_all.column_dimensions[get_column_letter(ci)].width = w
    ws_all.freeze_panes = "A2"

    col_idx    = {h: i + 1 for i, h in enumerate(_MULTI_HEADERS)}
    status_col   = col_idx["Status"]
    priority_col = col_idx["Priority"]
    date_cols    = {col_idx[h] for h in _MULTI_DATE_COLS}

    all_row = 1
    for proj_name, tasks in project_tasks:
        if not tasks:
            continue
        # Section header spanning all columns
        n_cols = len(_MULTI_HEADERS)
        ws_all.merge_cells(start_row=all_row, start_column=1, end_row=all_row, end_column=n_cols)
        sc = ws_all.cell(row=all_row, column=1, value=proj_name)
        sc.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        sc.fill = PatternFill(fill_type="solid", fgColor="2F5597")
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_all.row_dimensions[all_row].height = 18
        all_row += 1

        # Column headers
        for ci, hdr in enumerate(_MULTI_HEADERS, start=1):
            cell = ws_all.cell(row=all_row, column=ci, value=hdr)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _BORDER
        ws_all.row_dimensions[all_row].height = 18
        all_row += 1

        # Data rows
        notes_ci = col_idx["Notes"]
        notes_w_all = _MULTI_COL_WIDTHS[notes_ci - 1]
        cpl_all = max(1, int(notes_w_all * 1.1))

        for t in tasks:
            vals = _task_to_row(t)
            for ci, val in enumerate(vals, start=1):
                cell = ws_all.cell(row=all_row, column=ci, value=val)
                cell.font = _DATA_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=(ci == notes_ci))
                cell.border = _BORDER
                if ci in date_cols and val:
                    cell.number_format = "YYYY-MM-DD"
                if ci == status_col and val:
                    fhex = STATUS_FILLS.get(str(val).lower())
                    if fhex:
                        cell.fill = PatternFill(fill_type="solid", fgColor=fhex)
                if ci == priority_col and val:
                    fhex = PRIORITY_FILLS.get(str(val).lower())
                    if fhex:
                        cell.fill = PatternFill(fill_type="solid", fgColor=fhex)
            notes_val = str(vals[notes_ci - 1] or "")
            n_lines = max(1, -(-len(notes_val) // cpl_all))
            ws_all.row_dimensions[all_row].height = max(15, n_lines * 14)
            all_row += 1

        all_row += 1  # blank spacer

    if all_row == 1:
        ws_all.cell(row=1, column=1, value="No tasks found.")

    # ── Per-project sheets ────────────────────────────────────────────────────
    for proj_name, tasks in project_tasks:
        ws = wb.create_sheet(proj_name)
        _write_multi_sheet(ws, tasks)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"tasks_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-kanban")
async def export_tasks_kanban(
    project_ids: list[str] = Query(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export all tasks as a Kanban board layout grouped by date buckets."""
    from datetime import datetime as dt, date, timedelta

    # ── Collect tasks ────────────────────────────────────────────────────────
    all_tasks: list[Task] = []
    for pid in project_ids:
        result = await db.execute(
            select(Task).where(Task.project_id == pid).order_by(Task.due_date)
        )
        all_tasks.extend(result.scalars().all())

    # ── Date bucket helpers ───────────────────────────────────────────────────
    def _get_monday(d: date) -> date:
        return d - timedelta(days=d.weekday())

    def _get_bucket(due: date | None) -> str:
        if due is None:
            return "no_date"
        today = date.today()
        this_mon = _get_monday(today)
        next_mon = this_mon + timedelta(days=7)
        after_next = next_mon + timedelta(days=7)
        if due < today:
            return "past"
        if due == today:
            return "today"
        if this_mon <= due < next_mon:
            return "this_week"
        if next_mon <= due < after_next:
            return "next_week"
        return "later"

    # ── Group tasks ───────────────────────────────────────────────────────────
    BUCKET_ORDER = ["past", "today", "this_week", "next_week", "later", "no_date"]
    BUCKET_CONFIG = {
        "past":      ("Past Dates",      "C00000", "FFE0E0"),
        "today":     ("Today",           "375623", "C6EFCE"),
        "this_week": ("This Week",       "1F5C6B", "DDEBF7"),
        "next_week": ("Next Week",       "1F5C6B", "BDD7EE"),
        "later":     ("Later",           "7F6000", "FFEB9C"),
        "no_date":   ("Without a Date",  "595959", "F2F2F2"),
    }
    STATUS_ORDER = ["todo", "in_progress", "in_review", "done"]
    STATUS_LABELS = ["TODO", "IN PROGRESS", "IN REVIEW", "DONE"]
    CARD_PRIORITY_FILLS = {
        "critical": "FFC7CE",
        "high":     "FFEB9C",
        "medium":   "FFF2CC",
    }

    grouped: dict[str, list[Task]] = {b: [] for b in BUCKET_ORDER}
    for task in all_tasks:
        grouped[_get_bucket(task.due_date)].append(task)

    # ── Card text builder (strip HTML from notes) ────────────────────────────
    def _build_card(task: Task) -> str:
        return "\n".join([
            f"{task.external_id or '—'} # {task.title}",
            f"Type: {task.task_type or '—'}",
            f"Due: {task.due_date or '—'}",
            f"Notes: {_html_to_text(task.notes)[:80]}",
            f"Priority: {task.priority.value}",
            f"Approv: {'Yes' if task.approval else 'No'}",
        ])

    # ── Build workbook ────────────────────────────────────────────────────────
    thin = Side(style="thin", color="BFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kanban Board"

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 46

    row = 1

    # Title
    ws.merge_cells(f"A{row}:D{row}")
    tc = ws.cell(row=row, column=1, value="FELADAT STÁTUSZ")
    tc.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    tc.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26
    row += 1

    # Column headers
    for ci, lbl in enumerate(STATUS_LABELS, start=1):
        hc = ws.cell(row=row, column=ci, value=lbl)
        hc.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        hc.fill = PatternFill(fill_type="solid", fgColor="4472C4")
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border = cell_border
    ws.row_dimensions[row].height = 18
    row += 1

    # Sections
    for bucket_id in BUCKET_ORDER:
        tasks_in_bucket = grouped[bucket_id]
        if not tasks_in_bucket:
            continue

        label, text_color, bg_color = BUCKET_CONFIG[bucket_id]

        # Section header
        ws.merge_cells(f"A{row}:D{row}")
        sc = ws.cell(row=row, column=1, value=label)
        sc.font = Font(name="Calibri", bold=True, size=12, color=text_color)
        sc.fill = PatternFill(fill_type="solid", fgColor=bg_color)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        by_status: dict[str, list[Task]] = {s: [] for s in STATUS_ORDER}
        for task in tasks_in_bucket:
            by_status[task.status.value].append(task)

        max_cards = max(len(v) for v in by_status.values())

        for ri in range(max_cards):
            for ci, status in enumerate(STATUS_ORDER, start=1):
                col_tasks = by_status[status]
                if ri < len(col_tasks):
                    t = col_tasks[ri]
                    cc = ws.cell(row=row, column=ci, value=_build_card(t))
                    cc.alignment = Alignment(vertical="top", wrap_text=True)
                    cc.font = Font(name="Calibri", size=9)
                    cc.border = cell_border
                    prio_hex = CARD_PRIORITY_FILLS.get(t.priority.value)
                    if prio_hex:
                        cc.fill = PatternFill(fill_type="solid", fgColor=prio_hex)
                else:
                    ws.cell(row=row, column=ci).border = cell_border
            ws.row_dimensions[row].height = 75
            row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"kanban_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export")
async def export_tasks(
    project_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = select(Task)
    if project_id:
        q = q.where(Task.project_id == project_id)
    result = await db.execute(q.order_by(Task.created_at.desc()))
    tasks = result.scalars().all()

    rows = [
        {
            "Task ID": t.external_id or "",
            "Title": t.title,
            "Type": t.task_type or "",
            "Status": t.status.value,
            "Priority": t.priority.value,
            "Listing Date": t.listing_date,
            "Due Date": t.due_date,
            "Finishing Date": t.finishing_date,
            "Notes": t.notes or "",
            "Approval": t.approval,
        }
        for t in tasks
    ]

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tasks")
        _style_worksheet(writer.sheets["Tasks"])
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tasks.xlsx"},
    )


# ── HTML → plain text (preserves link URLs) ──────────────────────────────────

_RE_LINK  = re.compile(r'<a\b[^>]*\bhref=["\']([^"\']+)["\'][^>]*>(.*?)</a>', re.IGNORECASE | re.DOTALL)
_RE_BR    = re.compile(r'<br\s*/?>', re.IGNORECASE)
_RE_TAG   = re.compile(r'<[^>]+>')


def _html_to_text(html: str | None) -> str:
    """Convert HTML to plain text, preserving anchor link URLs as 'text (url)'."""
    if not html:
        return ""
    # Convert <a href="url">text</a>  →  "text (url)"
    text = _RE_LINK.sub(lambda m: f"{m.group(2).strip()} ({m.group(1).strip()})", html)
    # Convert <br> to newline
    text = _RE_BR.sub("\n", text)
    # Strip remaining tags
    text = _RE_TAG.sub("", text)
    # Decode common HTML entities
    text = (text
            .replace("&amp;", "&")
            .replace("&lt;", "<")
            .replace("&gt;", ">")
            .replace("&nbsp;", " ")
            .replace("&quot;", '"')
            .replace("&#39;", "'"))
    return text.strip()


# ── PDF helpers ───────────────────────────────────────────────────────────────


def _new_pdf(orientation: str = "L") -> "FPDF":
    if not _FPDF_AVAILABLE:
        raise HTTPException(status_code=503, detail="PDF library not available — rebuild the container.")
    pdf = FPDF(orientation=orientation, unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    _ensure_fonts(pdf)
    pdf.add_page()
    return pdf


_FONT_UNICODE = "/app/fonts/ArialUnicode.ttf"
_FONTS_LOADED = False


def _ensure_fonts(pdf: "FPDF") -> None:
    """Register Arial Unicode for full UTF-8 support; fall back to Helvetica if unavailable."""
    import os
    global _FONTS_LOADED
    if os.path.exists(_FONT_UNICODE):
        # Arial Unicode has no separate bold variant — register same file for both
        pdf.add_font("Unicode", style="",  fname=_FONT_UNICODE)
        pdf.add_font("Unicode", style="B", fname=_FONT_UNICODE)
        _FONTS_LOADED = True
    else:
        _FONTS_LOADED = False


def _pf(text: str, max_len: int = 500) -> str:
    """Prepare text for PDF: truncate. When Unicode font is loaded, passthrough; else latin-1 safe."""
    import unicodedata
    s = str(text or "")[:max_len]
    if not _FONTS_LOADED:
        s = unicodedata.normalize("NFKD", s).encode("latin-1", errors="replace").decode("latin-1")
    return s


def _set_font(pdf: "FPDF", style: str = "", size: int = 8) -> None:
    if _FONTS_LOADED:
        pdf.set_font("Unicode", style=style, size=size)
    else:
        import unicodedata
        # Helvetica fallback: normalize to latin-1
        pdf.set_font("Helvetica", style=style, size=size)


def _pdf_section(
    pdf: "FPDF",
    rows: list[dict],
    col_widths: list[float],
    headers: list[str],
    section_title: str | None = None,
    section_color: tuple[int, int, int] = (68, 114, 196),
    line_h: float = 4.5,
) -> None:
    """Append a section header row + table body with wrapped text, equal-height rows."""
    total_w = sum(col_widths)

    if section_title:
        pdf.set_fill_color(*section_color)
        pdf.set_text_color(255, 255, 255)
        _set_font(pdf, "B", 9)
        pdf.cell(total_w, 7, _pf(section_title), border=0, new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.set_text_color(0, 0, 0)

    # Column header row
    pdf.set_fill_color(220, 230, 242)
    _set_font(pdf, "B", 8)
    for w, lbl in zip(col_widths, headers):
        pdf.cell(w, 6, _pf(lbl), border=1, new_x="RIGHT", new_y="TOP", fill=True)
    pdf.ln()

    # Data rows — equal height per row (tallest cell determines height)
    _set_font(pdf, "", 8)
    chars_per_mm = 0.52   # approximate chars per mm at size 8, DejaVuSansCondensed

    for row in rows:
        vals = [_pf(str(row.get(key, "") or "")) for key in headers]

        # Calculate line count per cell → row height
        line_counts = []
        for w, val in zip(col_widths, vals):
            chars_per_line = max(1, int(w * chars_per_mm))
            n_lines = max(1, -(-len(val) // chars_per_line))
            line_counts.append(n_lines)
        row_h = max(line_counts) * line_h

        # Page break check
        if pdf.get_y() + row_h > pdf.h - 12:
            pdf.add_page()

        y_start = pdf.get_y()
        x_start = pdf.l_margin

        # Draw cell backgrounds + borders (full row_h) then write text
        pdf.set_fill_color(255, 255, 255)
        pdf.set_draw_color(180, 180, 180)
        x = x_start
        for w in col_widths:
            pdf.rect(x, y_start, w, row_h, style="FD")
            x += w

        # Write text inside each cell
        x = x_start
        for w, val in zip(col_widths, vals):
            pdf.set_xy(x + 0.8, y_start + 0.5)
            pdf.multi_cell(w - 1.0, line_h, val, border=0, max_line_height=line_h)
            x += w

        pdf.set_y(y_start + row_h)

    pdf.ln(3)


# ── Shared bucket helpers ─────────────────────────────────────────────────────

def _bucket_tasks(all_tasks: list[Task]):
    from datetime import date as date_t, timedelta

    today = date_t.today()
    this_monday = today - timedelta(days=today.weekday())
    next_monday = this_monday + timedelta(weeks=1)
    after_next = next_monday + timedelta(weeks=1)

    BUCKET_ORDER = ["past", "today", "this_week", "next_week", "later", "no_date"]
    BUCKET_LABELS = {
        "past":      "Past Dates",
        "today":     "Today",
        "this_week": "This Week",
        "next_week": "Next Week",
        "later":     "Later",
        "no_date":   "Without a Date",
    }
    BUCKET_COLORS = {
        "past":      (192, 80, 77),
        "today":     (84, 130, 53),
        "this_week": (31, 73, 125),
        "next_week": (23, 100, 120),
        "later":     (143, 121, 23),
        "no_date":   (119, 119, 119),
    }

    def _b(t: Task) -> str:
        d = t.due_date
        if not d:                           return "no_date"
        if d < today:                       return "past"
        if d == today:                      return "today"
        if this_monday <= d < next_monday:  return "this_week"
        if next_monday <= d < after_next:   return "next_week"
        return "later"

    grouped: dict[str, list[Task]] = {b: [] for b in BUCKET_ORDER}
    for t in all_tasks:
        grouped[_b(t)].append(t)

    return grouped, BUCKET_ORDER, BUCKET_LABELS, BUCKET_COLORS


def _task_row_dict(t: Task) -> dict:
    return {
        "Task ID":  t.external_id or "",
        "Title":    t.title,
        "Type":     t.task_type or "",
        "Status":   t.status.value.replace("_", " "),
        "Priority": t.priority.value,
        "Due Date": str(t.due_date) if t.due_date else "",
        "Notes":    _html_to_text(t.notes),
    }


# ── Timeline exports ─────────────────────────────────────────────────────────

TIMELINE_BUCKET_EXCEL = {
    "past":      ("Past Dates",      "C00000", "FFE0E0"),
    "today":     ("Today",           "375623", "C6EFCE"),
    "this_week": ("This Week",       "1F5C6B", "DDEBF7"),
    "next_week": ("Next Week",       "1F5C6B", "BDD7EE"),
    "later":     ("Later",           "7F6000", "FFEB9C"),
    "no_date":   ("Without a Date",  "595959", "F2F2F2"),
}

TIMELINE_COLUMNS = [
    ("Task ID",       "A", 14),
    ("Title",         "B", 42),
    ("Type",          "C", 16),
    ("Status",        "D", 16),
    ("Priority",      "E", 14),
    ("Listing Date",  "F", 16),
    ("Due Date",      "G", 16),
    ("Finishing Date","H", 18),
    ("Notes",         "I", 40),
    ("Approval",      "J", 12),
    ("Project",       "K", 24),
]

_thin_tl = Side(style="thin", color="BFBFBF")
_BORDER_TL = Border(left=_thin_tl, right=_thin_tl, top=_thin_tl, bottom=_thin_tl)


def _write_timeline_sheet(ws, tasks: list[Task], project_map: dict[str, str]) -> None:
    """Write a single timeline bucket sheet."""
    # Column widths  (TIMELINE_COLUMNS = [name, letter, width])
    for _, col_letter, width in TIMELINE_COLUMNS:
        ws.column_dimensions[col_letter].width = width

    # Header row
    headers = [c[0] for c in TIMELINE_COLUMNS]
    for ci, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_TL
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # Data rows  (Notes is col 9 / index 8 in TIMELINE_COLUMNS)
    _notes_ci = 9
    _notes_w  = TIMELINE_COLUMNS[8][2]
    _cpl      = max(1, int(_notes_w * 1.1))

    for ri, t in enumerate(tasks, start=2):
        vals = [
            t.external_id or "",
            t.title,
            t.task_type or "",
            t.status.value.replace("_", " "),
            t.priority.value,
            str(t.listing_date) if t.listing_date else "",
            str(t.due_date) if t.due_date else "",
            str(t.finishing_date) if t.finishing_date else "",
            _html_to_text(t.notes),
            "Yes" if t.approval else "No",
            project_map.get(t.project_id or "", ""),
        ]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == _notes_ci))
            cell.border = _BORDER_TL
        notes_val = vals[_notes_ci - 1]
        n_lines = max(1, -(-len(notes_val) // _cpl))
        ws.row_dimensions[ri].height = max(15, n_lines * 14)


@router.get("/export-timeline")
async def export_timeline(
    project_ids: list[str] = Query(default=[]),
    project_names: list[str] = Query(default=[]),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export Timeline view as Excel — one sheet per date bucket, table format."""
    from datetime import datetime as dt

    # Collect tasks across all projects
    all_tasks: list[Task] = []
    project_map: dict[str, str] = {}
    for i, pid in enumerate(project_ids):
        name = project_names[i] if i < len(project_names) else f"Project {i + 1}"
        project_map[pid] = name
        result = await db.execute(
            select(Task).where(Task.project_id == pid).order_by(Task.due_date)
        )
        all_tasks.extend(result.scalars().all())

    grouped, BUCKET_ORDER, _, _ = _bucket_tasks(all_tasks)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── "All Tasks" combined sheet ────────────────────────────────────────────
    ws_all = wb.create_sheet(title="All Tasks")
    ws_all.sheet_properties.tabColor = "1F4E79"
    for _, col_letter, width in TIMELINE_COLUMNS:
        ws_all.column_dimensions[col_letter].width = width

    col_count = len(TIMELINE_COLUMNS)
    all_row = 1
    headers = [c[0] for c in TIMELINE_COLUMNS]

    # Section color fills for the combined sheet
    BUCKET_SECTION_FILLS = {
        "past":      ("FFE0E0", "C00000"),
        "today":     ("C6EFCE", "375623"),
        "this_week": ("DDEBF7", "1F5C6B"),
        "next_week": ("BDD7EE", "1F5C6B"),
        "later":     ("FFEB9C", "7F6000"),
        "no_date":   ("F2F2F2", "595959"),
    }

    for b in BUCKET_ORDER:
        tasks = grouped[b]
        if not tasks:
            continue
        label, text_color, bg_color = TIMELINE_BUCKET_EXCEL[b]

        # Section header spanning all columns
        ws_all.merge_cells(start_row=all_row, start_column=1, end_row=all_row, end_column=col_count)
        sc = ws_all.cell(row=all_row, column=1, value=label)
        sc.font = Font(name="Calibri", bold=True, size=11, color=text_color)
        sc.fill = PatternFill(fill_type="solid", fgColor=bg_color)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_all.row_dimensions[all_row].height = 18
        all_row += 1

        # Column headers
        for ci, hdr in enumerate(headers, start=1):
            cell = ws_all.cell(row=all_row, column=ci, value=hdr)
            cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER_TL
        ws_all.row_dimensions[all_row].height = 15
        all_row += 1

        # Data rows  (Notes is at index 8, col 9)
        _notes_ci_tl = 9   # 1-based column index of Notes in TIMELINE_COLUMNS
        _notes_w_tl  = TIMELINE_COLUMNS[8][2]  # width in chars
        _cpl_tl      = max(1, int(_notes_w_tl * 1.1))

        for t in tasks:
            vals = [
                t.external_id or "",
                t.title,
                t.task_type or "",
                t.status.value.replace("_", " "),
                t.priority.value,
                str(t.listing_date) if t.listing_date else "",
                str(t.due_date) if t.due_date else "",
                str(t.finishing_date) if t.finishing_date else "",
                _html_to_text(t.notes),
                "Yes" if t.approval else "No",
                project_map.get(t.project_id or "", ""),
            ]
            for ci, val in enumerate(vals, start=1):
                cell = ws_all.cell(row=all_row, column=ci, value=val)
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=(ci == _notes_ci_tl))
                cell.border = _BORDER_TL
            notes_val = vals[_notes_ci_tl - 1]
            n_lines = max(1, -(-len(notes_val) // _cpl_tl))
            ws_all.row_dimensions[all_row].height = max(15, n_lines * 14)
            all_row += 1

        all_row += 1  # blank spacer between buckets

    if all_row == 1:
        ws_all.cell(row=1, column=1, value="No tasks found.")

    # ── Per-bucket sheets ─────────────────────────────────────────────────────
    for b in BUCKET_ORDER:
        tasks = grouped[b]
        if not tasks:
            continue
        label, _, _ = TIMELINE_BUCKET_EXCEL[b]
        ws = wb.create_sheet(title=label)
        ws.sheet_properties.tabColor = TIMELINE_BUCKET_EXCEL[b][1]
        _write_timeline_sheet(ws, tasks, project_map)

    if not wb.worksheets:
        ws = wb.create_sheet("No Tasks")
        ws.cell(row=1, column=1, value="No tasks found.")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"timeline_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-timeline-pdf")
async def export_timeline_pdf(
    project_ids: list[str] = Query(default=[]),
    project_names: list[str] = Query(default=[]),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export Timeline view as PDF — sections per date bucket."""
    from datetime import datetime as dt

    all_tasks: list[Task] = []
    project_map: dict[str, str] = {}
    for i, pid in enumerate(project_ids):
        name = project_names[i] if i < len(project_names) else f"Project {i + 1}"
        project_map[pid] = name
        result = await db.execute(
            select(Task).where(Task.project_id == pid).order_by(Task.due_date)
        )
        all_tasks.extend(result.scalars().all())

    grouped, BUCKET_ORDER, BUCKET_LABELS, BUCKET_COLORS = _bucket_tasks(all_tasks)

    pdf = _new_pdf()
    _set_font(pdf, "B", 14)
    pdf.cell(0, 10, "Timeline Export", border=0, align="C", new_x="LMARGIN", new_y="NEXT")
    _set_font(pdf, "", 9)
    pdf.cell(0, 6, dt.now().strftime("Generated: %Y-%m-%d %H:%M"), border=0, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    headers = ["Task ID", "Title", "Type", "Status", "Priority", "Due Date", "Project", "Notes"]
    col_widths = [22.0, 72.0, 24.0, 24.0, 20.0, 22.0, 30.0, 53.0]

    for b in BUCKET_ORDER:
        tasks = grouped[b]
        if not tasks:
            continue
        rows = []
        for t in tasks:
            r = _task_row_dict(t)
            r["Project"] = project_map.get(t.project_id or "", "")
            rows.append(r)
        _pdf_section(pdf, rows, col_widths, headers, section_title=BUCKET_LABELS[b], section_color=BUCKET_COLORS[b])

    output = io.BytesIO(pdf.output())
    filename = f"timeline_{dt.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-multi-pdf")
async def export_tasks_multi_pdf(
    project_ids: list[str] = Query(default=[]),
    project_names: list[str] = Query(default=[]),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export tasks for multiple projects as PDF, one section per project."""
    from datetime import datetime as dt

    pdf = _new_pdf()
    _set_font(pdf, "B", 14)
    pdf.cell(0, 10, "Task Export", border=0, align="C", new_x="LMARGIN", new_y="NEXT")
    _set_font(pdf, "", 9)
    pdf.cell(0, 6, dt.now().strftime("Generated: %Y-%m-%d %H:%M"), border=0, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    headers = ["Task ID", "Title", "Type", "Status", "Priority", "Due Date", "Notes"]
    col_widths = [22.0, 78.0, 26.0, 26.0, 22.0, 24.0, 69.0]

    for i, pid in enumerate(project_ids):
        result = await db.execute(
            select(Task).where(Task.project_id == pid).order_by(Task.created_at.desc())
        )
        tasks = result.scalars().all()
        rows = [
            {
                "Task ID": t.external_id or "",
                "Title": t.title,
                "Type": t.task_type or "",
                "Status": t.status.value.replace("_", " "),
                "Priority": t.priority.value,
                "Due Date": str(t.due_date) if t.due_date else "",
                "Notes": _html_to_text(t.notes),
            }
            for t in tasks
        ]
        section = project_names[i] if i < len(project_names) else f"Project {i + 1}"
        _pdf_section(pdf, rows, col_widths, headers, section_title=section)

    output = io.BytesIO(pdf.output())
    filename = f"tasks_{dt.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-kanban-pdf")
async def export_kanban_pdf(
    project_ids: list[str] = Query(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export tasks as a Kanban board PDF — 4 status columns, date-bucket sections."""
    from datetime import datetime as dt

    all_tasks: list[Task] = []
    for pid in project_ids:
        result = await db.execute(
            select(Task).where(Task.project_id == pid).order_by(Task.due_date)
        )
        all_tasks.extend(result.scalars().all())

    grouped, BUCKET_ORDER, BUCKET_LABELS, BUCKET_COLORS = _bucket_tasks(all_tasks)

    # ── Layout constants (landscape A4 = 297×210mm) ──────────────────────────
    MARGIN      = 8.0          # left/right margin
    PAGE_W      = 297.0
    USABLE_W    = PAGE_W - 2 * MARGIN
    COL_GAP     = 2.0
    COL_W       = (USABLE_W - 3 * COL_GAP) / 4   # ≈ 67.75mm
    CARD_H      = 26.0         # fixed card height
    CARD_PAD    = 1.5          # inner padding
    COL_HDR_H   = 7.0
    SEC_HDR_H   = 8.0
    STATUS_ORDER = ["todo", "in_progress", "in_review", "done"]
    STATUS_LABELS_PDF = ["TODO", "IN PROGRESS", "IN REVIEW", "DONE"]
    STATUS_FILLS = {          # (R, G, B) header fill
        "todo":        (100, 100, 110),
        "in_progress": (30, 100, 180),
        "in_review":   (160, 130, 20),
        "done":        (40, 130, 60),
    }
    PRIO_BG = {
        "critical": (255, 199, 206),
        "high":     (255, 235, 156),
        "medium":   (255, 255, 255),
        "low":      (242, 242, 242),
    }

    def col_x(ci: int) -> float:
        return MARGIN + ci * (COL_W + COL_GAP)

    pdf = _new_pdf("L")
    pdf.set_margins(MARGIN, 8, MARGIN)
    pdf.set_auto_page_break(auto=True, margin=10)

    # ── Title ─────────────────────────────────────────────────────────────────
    _set_font(pdf, "B", 14)
    pdf.cell(0, 10, "KANBAN BOARD", border=0, align="C", new_x="LMARGIN", new_y="NEXT")
    _set_font(pdf, "", 8)
    pdf.cell(0, 5, dt.now().strftime("Generated: %Y-%m-%d %H:%M"), border=0, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    for b in BUCKET_ORDER:
        tasks_in_bucket = grouped[b]
        if not tasks_in_bucket:
            continue

        by_status: dict[str, list[Task]] = {s: [] for s in STATUS_ORDER}
        for t in tasks_in_bucket:
            by_status[t.status.value].append(t)
        max_cards = max(len(v) for v in by_status.values())

        # How many rows this section needs (header + col headers + cards)
        section_h = SEC_HDR_H + COL_HDR_H + max_cards * CARD_H + 4
        if pdf.get_y() + section_h > pdf.h - 14:
            pdf.add_page()

        y = pdf.get_y()

        # Bucket section header (full width)
        r, g, b_col = BUCKET_COLORS[b]
        pdf.set_fill_color(r, g, b_col)
        pdf.set_text_color(255, 255, 255)
        _set_font(pdf, "B", 10)
        pdf.set_xy(MARGIN, y)
        pdf.cell(USABLE_W, SEC_HDR_H, _pf(BUCKET_LABELS[b]), border=0, fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        y += SEC_HDR_H

        # Column headers
        for ci, (status, lbl) in enumerate(zip(STATUS_ORDER, STATUS_LABELS_PDF)):
            sr, sg, sb = STATUS_FILLS[status]
            pdf.set_fill_color(sr, sg, sb)
            pdf.set_text_color(255, 255, 255)
            _set_font(pdf, "B", 9)
            pdf.set_xy(col_x(ci), y)
            pdf.cell(COL_W, COL_HDR_H, lbl, border=1, fill=True, align="C")
        pdf.set_text_color(0, 0, 0)
        y += COL_HDR_H

        # Card rows
        for ri in range(max_cards):
            # Check page break before each row
            if y + CARD_H > pdf.h - 10:
                pdf.add_page()
                y = pdf.get_y()

            for ci, status in enumerate(STATUS_ORDER):
                col_tasks = by_status[status]
                cx = col_x(ci)
                if ri < len(col_tasks):
                    t = col_tasks[ri]
                    bg = PRIO_BG.get(t.priority.value, (255, 255, 255))
                    pdf.set_fill_color(*bg)
                    # Card border + fill
                    pdf.set_draw_color(180, 180, 180)
                    pdf.rect(cx, y, COL_W, CARD_H, style="FD")
                    # Card text lines
                    _set_font(pdf, "B", 7)
                    pdf.set_xy(cx + CARD_PAD, y + CARD_PAD)
                    id_title = _pf(f"{t.external_id or '-'} {t.title}", 44)
                    pdf.cell(COL_W - 2 * CARD_PAD, 4, id_title)
                    _set_font(pdf, "", 7)
                    lines = [
                        _pf(f"Type: {t.task_type or '-'}", 38),
                        _pf(f"Due:  {t.due_date or '-'}", 38),
                        _pf(f"Prio: {t.priority.value}  Appr: {'Yes' if t.approval else 'No'}", 38),
                        _pf(f"Notes: {_html_to_text(t.notes)}", 44),
                    ]
                    for li, line in enumerate(lines):
                        pdf.set_xy(cx + CARD_PAD, y + CARD_PAD + 4 + li * 4.5)
                        pdf.cell(COL_W - 2 * CARD_PAD, 4, line)
                else:
                    # Empty cell
                    pdf.set_fill_color(248, 248, 250)
                    pdf.set_draw_color(210, 210, 210)
                    pdf.rect(cx, y, COL_W, CARD_H, style="FD")

            y += CARD_H

        pdf.set_y(y + 4)  # gap after section

    output = io.BytesIO(pdf.output())
    filename = f"kanban_{dt.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
